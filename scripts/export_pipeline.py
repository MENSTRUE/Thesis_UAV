"""Rebuild workbook summary report dari reports/ (adaptasi
scripts/export_summary.py pada project drone_e99_face_recognition).

Membaca:
- reports/runs/manifest_*.json        -> sheet Experiments (1 baris per run)
- reports/sessions/**/segment_stats_*.csv, activity_stats_*, identity_stats_*,
  accuracy_*, confusion_*, per_second_*.csv -> sheet tabel.

Baris run lama dipertahankan (tidak ditimpa) saat regenerate.
"""
import csv
import json
import sys
from glob import glob
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

RUNS_DIR = ROOT / "reports" / "runs"
SESSIONS_DIR = ROOT / "reports" / "sessions"
OUT = ROOT / "reports" / "summary" / "summary_pipeline.xlsx"

EXPERIMENT_HEADERS = [
    "timestamp", "session_id", "source", "profile", "n_segments",
    "n_frames", "throughput_fps", "labels",
    "detector_conf", "pose_conf", "pose_interval", "face_interval",
    "max_people", "face_enabled", "face_threshold", "liveness_threshold",
    "git_commit", "manifest",
]

STATS_SHEETS = [
    ("SegmentStats", "segment_stats_*.csv"),
    ("PerActivity", "activity_stats_*.csv"),
    ("PerIdentity", "identity_stats_*.csv"),
    ("Accuracy", "accuracy_*.csv"),
    ("ConfusionMatrix", "confusion_*.csv"),
    ("PerSecond", "per_second_*.csv"),
]


def _load_json(path):
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        print(f"[WARN] Manifest korup, dilewati: {path}")
        return None


def load_manifests():
    out = []
    for path in sorted(glob(str(RUNS_DIR / "manifest_*.json"))):
        data = _load_json(path)
        if data:
            out.append(data)
    return out


def merge_csv(pattern):
    """Gabung semua CSV yang cocok jadi list dict, tambahkan session_id."""
    rows = []
    for path in sorted(glob(pattern)):
        session = Path(path).parent.name
        with open(path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                row = {k: (v if v != "" else None) for k, v in row.items()}
                rows.append({"session_id": session, **row})
    return rows


def write_table(wb, title, headers, rows, best_cols=None):
    ws = wb.create_sheet(title)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    for r, row in enumerate(rows, 2):
        for c, h in enumerate(headers):
            ws.cell(row=r, column=c + 1, value=row.get(h))
    if best_cols:
        for col, direction in best_cols.items():
            if col not in headers:
                continue
            ci = headers.index(col) + 1
            vals = [ws.cell(row=r, column=ci).value for r in range(2, len(rows) + 2)]
            vals = [v for v in vals if isinstance(v, (int, float))]
            if not vals:
                continue
            target = max(vals) if direction == "max" else min(vals)
            fill = PatternFill("solid", fgColor="C6EFCE")
            for r in range(2, len(rows) + 2):
                if ws.cell(row=r, column=ci).value == target:
                    ws.cell(row=r, column=ci).fill = fill
    for c in range(1, len(headers) + 1):
        width = len(str(headers[c - 1]))
        for r in range(2, len(rows) + 2):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                width = max(width, len(str(val)))
        ws.column_dimensions[get_column_letter(c)].width = min(width + 2, 40)
    ws.freeze_panes = "A2"
    return ws


def experiment_rows(manifests):
    rows = []
    for m in manifests:
        cfg = m.get("config", {})
        env = m.get("env", {})
        segments = m.get("segments", [])
        n_frames = sum(s.get("frames", 0) for s in segments)
        fps_vals = [s.get("throughput_fps") for s in segments if s.get("throughput_fps")]
        rows.append({
            "timestamp": m.get("timestamp"),
            "session_id": m.get("session_id"),
            "source": m.get("source"),
            "profile": m.get("profile"),
            "n_segments": len(segments),
            "n_frames": n_frames,
            "throughput_fps": round(sum(fps_vals) / len(fps_vals), 2) if fps_vals else None,
            "labels": m.get("labels"),
            "detector_conf": cfg.get("detector_conf"),
            "pose_conf": cfg.get("pose_conf"),
            "pose_interval": cfg.get("pose_interval"),
            "face_interval": cfg.get("face_interval"),
            "max_people": cfg.get("max_people"),
            "face_enabled": cfg.get("face_enabled"),
            "face_threshold": cfg.get("face_threshold"),
            "liveness_threshold": cfg.get("liveness_threshold"),
            "git_commit": env.get("git_commit"),
            "manifest": m.get("run_id"),
        })
    return rows


def _load_existing_log():
    """Baca baris ExperimentLog dari workbook lama supaya tidak hilang."""
    if not OUT.exists():
        return []
    try:
        wb = load_workbook(OUT, read_only=True)
    except Exception:
        return []
    if "ExperimentLog" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["ExperimentLog"]
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    wb.close()
    return rows


def merge_unique_by_session(existing, new_rows):
    seen = {r.get("session_id") for r in existing}
    for row in new_rows:
        if row.get("session_id") and row["session_id"] in seen:
            continue
        existing.append(row)
        seen.add(row.get("session_id"))
    return existing


def build():
    manifests = load_manifests()
    new_rows = experiment_rows(manifests)
    exp_rows = merge_unique_by_session(_load_existing_log(), new_rows)

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    write_table(wb, "ExperimentLog", EXPERIMENT_HEADERS, exp_rows,
                best_cols={"throughput_fps": "max"})

    for sheet, pattern in STATS_SHEETS:
        rows = merge_csv(str(SESSIONS_DIR / "**" / pattern))
        if not rows and sheet not in ("SegmentStats",):
            continue
        if rows:
            headers = list(rows[0].keys())
            write_table(wb, sheet, headers, rows)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"[OK] {OUT} ({len(manifests)} run, {len(exp_rows)} baris experiment log)")


def main():
    build()


if __name__ == "__main__":
    main()